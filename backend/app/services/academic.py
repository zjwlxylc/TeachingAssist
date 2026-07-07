import json
import csv
import math
from datetime import date, datetime
from io import BytesIO
from io import StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import AppError
from app.db.session import get_connection
from app.services import ai as ai_service


MAX_EXCEL_SIZE = 10 * 1024 * 1024
STANDARD_FIELDS = {
    "student_id": "学号",
    "name": "姓名",
    "class_name": "班级",
    "major": "专业",
    "college": "学院",
    "grade": "年级",
}
REQUIRED_FIELDS = {"student_id", "name", "class_name"}
MAX_NAME_LENGTH = 100


def _row_to_dict(row: Any) -> dict[str, Any]:
    return dict(row)


def _normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\ufeff", "").strip()
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return f"{value:g}".strip()
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _normalize_import_field(field: str, value: Any) -> str:
    text = _normalize_cell_value(value)
    if field == "student_id" and text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return " ".join(text.split())


def _looks_corrupted_text(value: str) -> bool:
    text = value.strip()
    if not text:
        return False
    if "\ufffd" in text:
        return True
    question_count = text.count("?")
    visible_count = sum(1 for char in text if not char.isspace())
    return question_count >= 2 and visible_count > 0 and question_count / visible_count >= 0.3


def _validate_display_name(value: str, label: str, code_prefix: str) -> str:
    normalized = " ".join(value.strip().split())
    if not normalized:
        raise AppError(f"{label}不能为空", code=f"{code_prefix}_REQUIRED")
    if len(normalized) > MAX_NAME_LENGTH:
        raise AppError(f"{label}不能超过 {MAX_NAME_LENGTH} 个字符", code=f"{code_prefix}_TOO_LONG")
    if _looks_corrupted_text(normalized):
        raise AppError(f"{label}疑似乱码，请重新输入", code=f"{code_prefix}_CORRUPTED")
    return normalized


def list_courses() -> list[dict[str, Any]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT c.id, c.name, c.teacher_id, c.teacher_name, c.created_at, c.updated_at,
                   COUNT(DISTINCT cc.class_id) AS class_count,
                   COUNT(DISTINCT st.id) AS student_count
            FROM courses c
            LEFT JOIN course_classes cc ON cc.course_id = c.id
            LEFT JOIN course_students cs ON cs.course_id = c.id
            LEFT JOIN students st ON st.id = cs.student_id AND st.is_active = 1
            GROUP BY c.id
            ORDER BY c.updated_at DESC, c.id DESC
            """
        ).fetchall()
    return [_row_to_dict(row) for row in rows]


def create_course(name: str, teacher_id: int | None, teacher_name: str | None) -> dict[str, Any]:
    name = _validate_display_name(name, "课程名称", "COURSE_NAME")
    with get_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO courses(name, teacher_id, teacher_name)
            VALUES (?, ?, ?)
            """,
            (name, teacher_id, teacher_name.strip() if teacher_name else None),
        )
        row = connection.execute("SELECT * FROM courses WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return _row_to_dict(row)


def list_classes(course_id: int | None = None) -> list[dict[str, Any]]:
    if course_id:
        sql = """
            SELECT cl.id, cl.name, cl.created_at, cl.updated_at,
                   COUNT(DISTINCT s.id) AS student_count
            FROM classes cl
            JOIN course_classes cc ON cc.class_id = cl.id
            LEFT JOIN students s ON s.class_id = cl.id AND s.is_active = 1
            WHERE cc.course_id = ?
            GROUP BY cl.id
            ORDER BY cl.name
        """
        params: tuple[Any, ...] = (course_id,)
    else:
        sql = """
            SELECT cl.id, cl.name, cl.created_at, cl.updated_at,
                   COUNT(DISTINCT s.id) AS student_count
            FROM classes cl
            LEFT JOIN students s ON s.class_id = cl.id AND s.is_active = 1
            GROUP BY cl.id
            ORDER BY cl.name
        """
        params = ()
    with get_connection() as connection:
        rows = connection.execute(sql, params).fetchall()
    return [_row_to_dict(row) for row in rows]


def create_class(name: str) -> dict[str, Any]:
    name = _validate_display_name(name, "班级名称", "CLASS_NAME")
    with get_connection() as connection:
        connection.execute(
            """
            INSERT INTO classes(name)
            VALUES (?)
            ON CONFLICT(name) DO UPDATE SET updated_at = datetime('now')
            """,
            (name,),
        )
        row = connection.execute("SELECT * FROM classes WHERE name = ?", (name,)).fetchone()
    return _row_to_dict(row)


def link_course_class(course_id: int, class_id: int) -> dict[str, Any]:
    with get_connection() as connection:
        course = connection.execute("SELECT id FROM courses WHERE id = ?", (course_id,)).fetchone()
        klass = connection.execute("SELECT id FROM classes WHERE id = ?", (class_id,)).fetchone()
        if course is None:
            raise AppError("课程不存在", code="COURSE_NOT_FOUND", status_code=404)
        if klass is None:
            raise AppError("班级不存在", code="CLASS_NOT_FOUND", status_code=404)
        connection.execute(
            "INSERT OR IGNORE INTO course_classes(course_id, class_id) VALUES (?, ?)",
            (course_id, class_id),
        )
        row = connection.execute(
            """
            SELECT cc.id, cc.course_id, cc.class_id, c.name AS course_name, cl.name AS class_name, cc.created_at
            FROM course_classes cc
            JOIN courses c ON c.id = cc.course_id
            JOIN classes cl ON cl.id = cc.class_id
            WHERE cc.course_id = ? AND cc.class_id = ?
            """,
            (course_id, class_id),
        ).fetchone()
    return _row_to_dict(row)


def list_sessions(course_id: int | None = None) -> list[dict[str, Any]]:
    from app.services.classroom import refresh_session_statuses

    refresh_session_statuses()
    params: tuple[Any, ...] = ()
    where = ""
    if course_id:
        where = "WHERE s.course_id = ?"
        params = (course_id,)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT s.*, c.name AS course_name, cl.name AS class_name,
                   COUNT(st.id) AS roster_count
            FROM classroom_sessions s
            JOIN courses c ON c.id = s.course_id
            JOIN classes cl ON cl.id = s.class_id
            LEFT JOIN course_students cs ON cs.course_id = s.course_id AND cs.class_id = s.class_id
            LEFT JOIN students st ON st.id = cs.student_id AND st.is_active = 1
            {where}
            GROUP BY s.id
            ORDER BY COALESCE(s.start_time, s.created_at) DESC, s.id DESC
            """,
            params,
        ).fetchall()
    return [_row_to_dict(row) for row in rows]


def create_session(payload: dict[str, Any]) -> dict[str, Any]:
    course_id = int(payload["course_id"])
    class_id = int(payload["class_id"])
    title = _validate_display_name(str(payload["title"]), "课堂标题", "SESSION_TITLE")
    session_no = int(payload["session_no"])
    if session_no <= 0:
        raise AppError("课次必须大于 0", code="SESSION_NO_INVALID")

    with get_connection() as connection:
        linked = connection.execute(
            "SELECT id FROM course_classes WHERE course_id = ? AND class_id = ?",
            (course_id, class_id),
        ).fetchone()
        if linked is None:
            raise AppError("请先关联课程和班级", code="COURSE_CLASS_NOT_LINKED")
        try:
            cursor = connection.execute(
                """
                INSERT INTO classroom_sessions(
                    course_id, class_id, title, session_no, start_time, end_time,
                    is_makeup, schedule_note
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    course_id,
                    class_id,
                    title,
                    session_no,
                    payload.get("start_time"),
                    payload.get("end_time"),
                    1 if payload.get("is_makeup") else 0,
                    payload.get("schedule_note"),
                ),
            )
        except Exception as exc:
            if "UNIQUE" in str(exc):
                raise AppError("同一课程班级下课次已存在", code="SESSION_NO_EXISTS") from exc
            raise
        row = connection.execute("SELECT * FROM classroom_sessions WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return _row_to_dict(row)


def parse_excel_upload(file_name: str, file_size: int, content: bytes) -> dict[str, Any]:
    suffix = Path(file_name).suffix.lower()
    if suffix not in {".xls", ".xlsx"}:
        raise AppError("仅支持 .xls/.xlsx 格式", code="EXCEL_FORMAT_UNSUPPORTED")
    if file_size > MAX_EXCEL_SIZE:
        raise AppError("文件超过 10MB", code="EXCEL_FILE_TOO_LARGE")
    if suffix == ".xls":
        raise AppError("当前环境暂不支持旧版 .xls，请另存为 .xlsx 后上传", code="XLS_NOT_SUPPORTED")

    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise AppError("Excel 内容为空", code="EXCEL_EMPTY")

    # 跳过开头全空的行，定位真正的表头（兼容首行标题/留白行）
    header_index = 0
    while header_index < len(rows):
        candidate = [_normalize_cell_value(value) for value in rows[header_index]]
        if any(candidate):
            headers = candidate
            break
        header_index += 1
    else:
        raise AppError("Excel 表头为空", code="EXCEL_HEADER_EMPTY")

    data_rows: list[dict[str, str]] = []
    # 数据行从表头下一行开始；Excel 行号从 1 计，表头在第 header_index+1 行，故数据起始行号 header_index+2
    for row_index, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        item: dict[str, str] = {"__row_number": str(row_index)}
        has_value = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            normalized = _normalize_cell_value(value)
            if normalized:
                has_value = True
            item[header] = normalized
        if has_value:
            data_rows.append(item)

    with get_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO student_import_jobs(file_name, file_size, headers_json, sample_rows_json, rows_json)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                file_name,
                file_size,
                json.dumps(headers, ensure_ascii=False),
                json.dumps(data_rows[:5], ensure_ascii=False),
                json.dumps(data_rows, ensure_ascii=False),
            ),
        )
        job_id = cursor.lastrowid

    return {
        "job_id": job_id,
        "file_name": file_name,
        "file_size": file_size,
        "headers": headers,
        "sample_rows": data_rows[:5],
        "total_rows": len(data_rows),
        "standard_fields": STANDARD_FIELDS,
        "required_fields": sorted(REQUIRED_FIELDS),
    }


def _load_import_job(job_id: int) -> dict[str, Any]:
    with get_connection() as connection:
        row = connection.execute("SELECT * FROM student_import_jobs WHERE id = ?", (job_id,)).fetchone()
    if row is None:
        raise AppError("导入任务不存在", code="IMPORT_JOB_NOT_FOUND", status_code=404)
    job = _row_to_dict(row)
    job["headers"] = json.loads(str(job["headers_json"]))
    job["rows"] = json.loads(str(job["rows_json"]))
    return job


def preview_import(job_id: int, mapping: dict[str, str]) -> dict[str, Any]:
    preview = _build_import_preview(_load_import_job(job_id), mapping)
    preview.pop("all_rows", None)
    return preview


def suggest_import_mapping(job_id: int) -> dict[str, Any]:
    job = _load_import_job(job_id)
    headers = job["headers"]
    sample_rows = job["rows"][:5]
    prompt = (
        "请根据 Excel 表头和前 5 行样例，将原始列名映射到标准字段。"
        "仅返回 JSON，格式为 {\"mapping\":{\"原始列名\":\"student_id|name|class_name|major|college|grade\"},"
        "\"confidence\":0.0,\"notes\":\"...\"}。\n"
        f"标准字段：{json.dumps(STANDARD_FIELDS, ensure_ascii=False)}\n"
        f"表头：{json.dumps(headers, ensure_ascii=False)}\n"
        f"样例：{json.dumps(sample_rows, ensure_ascii=False)}"
    )
    mapping: dict[str, str] | None = None
    ai_error: str | None = None
    try:
        generated = ai_service.generate_structured_json(
            "excel_mapping",
            prompt,
            source_type="student_import_job",
            source_id=job_id,
        )
        raw_mapping = generated.get("mapping") if isinstance(generated, dict) else None
        if isinstance(raw_mapping, dict):
            mapping = {
                str(source): str(target)
                for source, target in raw_mapping.items()
                if source in headers and str(target) in STANDARD_FIELDS
            }
    except Exception as exc:
        ai_error = str(exc)
        ai_service.record_failure_task(
            "excel_mapping",
            "student_import_job",
            job_id,
            ai_error,
            {"headers": headers, "sample_rows": sample_rows},
        )

    if not mapping:
        mapping = _heuristic_mapping(headers)
        return {
            "mode": "manual_fallback",
            "message": "AI 不可用，请手动映射" if ai_error else "已按表头生成本地映射建议",
            "mapping": mapping,
            "standard_fields": STANDARD_FIELDS,
        }

    return {
        "mode": "ai_suggested",
        "message": "AI 字段映射建议已生成",
        "mapping": mapping,
        "standard_fields": STANDARD_FIELDS,
    }


def _heuristic_mapping(headers: list[str]) -> dict[str, str]:
    aliases = {
        "student_id": ["学号", "学生编号", "student_id", "student no", "id"],
        "name": ["姓名", "学生姓名", "name"],
        "class_name": ["班级", "班级名称", "class", "class_name"],
        "major": ["专业", "major"],
        "college": ["学院", "院系", "college"],
        "grade": ["年级", "grade"],
    }
    mapping: dict[str, str] = {}
    for header in headers:
        normalized = header.strip().lower()
        for field, names in aliases.items():
            if any(alias.lower() in normalized or normalized in alias.lower() for alias in names):
                mapping[header] = field
                break
    return mapping


def _build_import_preview(job: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    reverse_mapping = {standard: source for source, standard in mapping.items() if standard}
    missing = sorted(REQUIRED_FIELDS - set(reverse_mapping))
    if missing:
        raise AppError("必填字段未完成映射：" + "、".join(STANDARD_FIELDS[item] for item in missing), code="IMPORT_MAPPING_REQUIRED")

    # Batch-check existing student IDs to avoid N+1 connection overhead
    all_student_numbers: list[str] = []
    for raw in job["rows"]:
        normalized = {
            field: _normalize_import_field(field, raw.get(source, ""))
            for field, source in reverse_mapping.items()
        }
        student_number = normalized.get("student_id", "")
        if student_number:
            all_student_numbers.append(student_number)

    existing_student_ids: set[str] = set()
    if all_student_numbers:
        with get_connection() as connection:
            placeholders = ",".join("?" for _ in all_student_numbers)
            existing_rows = connection.execute(
                f"SELECT DISTINCT student_id FROM students WHERE student_id IN ({placeholders})",
                tuple(all_student_numbers),
            ).fetchall()
        existing_student_ids = {row["student_id"] for row in existing_rows}

    seen: dict[str, int] = {}
    rows: list[dict[str, Any]] = []
    errors = 0
    warnings = 0
    for raw in job["rows"]:
        normalized = {
            field: _normalize_import_field(field, raw.get(source, ""))
            for field, source in reverse_mapping.items()
        }
        row_errors: list[str] = []
        row_warnings: list[str] = []
        row_number = int(raw.get("__row_number", "0"))
        if not normalized.get("student_id"):
            row_errors.append("学号为空")
        if not normalized.get("name"):
            row_errors.append("姓名为空")
        if not normalized.get("class_name"):
            row_errors.append("班级为空")
        for field in ("name", "class_name"):
            if _looks_corrupted_text(normalized.get(field, "")):
                row_errors.append(f"{STANDARD_FIELDS[field]}疑似乱码")
        student_number = normalized.get("student_id", "")
        if student_number:
            if student_number in seen:
                row_errors.append(f"学号重复，首次出现在第 {seen[student_number]} 行")
            else:
                seen[student_number] = row_number
        if student_number and student_number in existing_student_ids:
            row_warnings.append("学号已存在，确认导入时会更新学生信息")
        errors += len(row_errors)
        warnings += len(row_warnings)
        rows.append(
            {
                "row_number": row_number,
                "data": normalized,
                "errors": row_errors,
                "warnings": row_warnings,
                "valid": not row_errors,
            }
        )

    return {
        "job_id": job["id"],
        "total_rows": len(rows),
        "valid_rows": sum(1 for row in rows if row["valid"]),
        "error_count": errors,
        "warning_count": warnings,
        "all_rows": rows,
        "rows": rows[:50],
    }


def confirm_import(
    job_id: int,
    course_id: int,
    mapping: dict[str, str],
    import_valid_only: bool = True,
    duplicate_strategy: str = "merge",
) -> dict[str, Any]:
    if duplicate_strategy not in {"overwrite", "skip", "merge"}:
        raise AppError("重复学号处理策略不支持", code="IMPORT_DUPLICATE_STRATEGY_INVALID")
    job = _load_import_job(job_id)
    preview = _build_import_preview(job, mapping)
    if preview["error_count"] and not import_valid_only:
        raise AppError("存在错误数据，不能全部导入", code="IMPORT_HAS_ERRORS")

    imported = 0
    updated = 0
    skipped = 0
    failed = 0
    report_rows: list[dict[str, Any]] = []
    with get_connection() as connection:
        course = connection.execute("SELECT id FROM courses WHERE id = ?", (course_id,)).fetchone()
        if course is None:
            raise AppError("课程不存在", code="COURSE_NOT_FOUND", status_code=404)

        for row in preview["all_rows"]:
            if not row["valid"]:
                skipped += 1
                report_rows.append(
                    {
                        "row_number": row["row_number"],
                        "student_id": row["data"].get("student_id", ""),
                        "reason": "；".join(row["errors"]),
                    }
                )
                continue
            data = row["data"]
            try:
                connection.execute(
                    """
                    INSERT INTO classes(name)
                    VALUES (?)
                    ON CONFLICT(name) DO UPDATE SET updated_at = datetime('now')
                    """,
                    (data["class_name"],),
                )
                class_row = connection.execute("SELECT id FROM classes WHERE name = ?", (data["class_name"],)).fetchone()
                class_id = int(class_row["id"])
                connection.execute(
                    "INSERT OR IGNORE INTO course_classes(course_id, class_id) VALUES (?, ?)",
                    (course_id, class_id),
                )
                existing_student = connection.execute(
                    "SELECT * FROM students WHERE student_id = ?",
                    (data["student_id"],),
                ).fetchone()
                if existing_student and duplicate_strategy == "skip":
                    skipped += 1
                    continue
                elif existing_student and duplicate_strategy == "merge":
                    connection.execute(
                        """
                        UPDATE students
                        SET name = COALESCE(NULLIF(?, ''), name),
                            class_id = ?,
                            major = COALESCE(NULLIF(?, ''), major),
                            college = COALESCE(NULLIF(?, ''), college),
                            grade = COALESCE(NULLIF(?, ''), grade),
                            is_active = 1,
                            updated_at = datetime('now')
                        WHERE student_id = ?
                        """,
                        (
                            data.get("name"),
                            class_id,
                            data.get("major"),
                            data.get("college"),
                            data.get("grade"),
                            data["student_id"],
                        ),
                    )
                    updated += 1
                else:
                    connection.execute(
                        """
                        INSERT INTO students(student_id, name, class_id, major, college, grade)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ON CONFLICT(student_id) DO UPDATE SET
                            name = excluded.name,
                            class_id = excluded.class_id,
                            major = excluded.major,
                            college = excluded.college,
                            grade = excluded.grade,
                            is_active = 1,
                            updated_at = datetime('now')
                        """,
                        (
                            data["student_id"],
                            data["name"],
                            class_id,
                            data.get("major"),
                            data.get("college"),
                            data.get("grade"),
                        ),
                    )
                    if existing_student:
                        updated += 1
                    else:
                        imported += 1
                student = connection.execute(
                    "SELECT id FROM students WHERE student_id = ?",
                    (data["student_id"],),
                ).fetchone()
                connection.execute(
                    """
                    INSERT INTO course_students(course_id, class_id, student_id)
                    VALUES (?, ?, ?)
                    ON CONFLICT(course_id, student_id) DO UPDATE SET
                        class_id = excluded.class_id
                    """,
                    (course_id, class_id, int(student["id"])),
                )
            except Exception:
                failed += 1
                report_rows.append(
                    {
                        "row_number": row["row_number"],
                        "student_id": data.get("student_id", ""),
                        "reason": "写入数据库失败",
                    }
                )
        if report_rows:
            connection.execute(
                """
                INSERT INTO student_import_reports(job_id, report_type, rows_json)
                VALUES (?, 'error', ?)
                """,
                (job_id, json.dumps(report_rows, ensure_ascii=False)),
            )
    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "failed": failed,
        "total": preview["total_rows"],
        "duplicate_strategy": duplicate_strategy,
        "error_report_available": bool(report_rows),
    }


def list_students(course_id: int | None = None, class_id: int | None = None, include_inactive: bool = False) -> list[dict[str, Any]]:
    params: list[Any] = []
    joins = ""
    where = ["1 = 1"] if include_inactive else ["s.is_active = 1"]
    if course_id:
        joins = "JOIN course_students cs ON cs.student_id = s.id"
        where.append("cs.course_id = ?")
        params.append(course_id)
    if class_id:
        where.append("s.class_id = ?")
        params.append(class_id)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT s.id, s.student_id, s.name, s.major, s.college, s.grade, s.is_active,
                   s.class_id, cl.name AS class_name, s.created_at, s.updated_at
            FROM students s
            JOIN classes cl ON cl.id = s.class_id
            {joins}
            WHERE {' AND '.join(where)}
            ORDER BY cl.name, s.student_id
            LIMIT 500
            """,
            tuple(params),
        ).fetchall()
    return [_row_to_dict(row) for row in rows]


def set_student_active(student_pk: int, is_active: bool) -> dict[str, Any]:
    with get_connection() as connection:
        row = connection.execute("SELECT * FROM students WHERE id = ?", (student_pk,)).fetchone()
        if row is None:
            raise AppError("学生不存在", code="STUDENT_NOT_FOUND", status_code=404)
        connection.execute(
            "UPDATE students SET is_active = ?, updated_at = datetime('now') WHERE id = ?",
            (1 if is_active else 0, student_pk),
        )
        updated = connection.execute(
            """
            SELECT s.id, s.student_id, s.name, s.major, s.college, s.grade, s.is_active,
                   s.class_id, cl.name AS class_name, s.created_at, s.updated_at
            FROM students s
            JOIN classes cl ON cl.id = s.class_id
            WHERE s.id = ?
            """,
            (student_pk,),
        ).fetchone()
    return _row_to_dict(updated)


def export_import_errors(job_id: int) -> dict[str, Any]:
    job = _load_import_job(job_id)
    preview_rows = _build_import_preview(job, _heuristic_mapping(job["headers"]))["all_rows"]
    report_rows = [
        {
            "row_number": row["row_number"],
            "student_id": row["data"].get("student_id", ""),
            "reason": "；".join(row["errors"] or row["warnings"]),
        }
        for row in preview_rows
        if row["errors"] or row["warnings"]
    ]
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=["row_number", "student_id", "reason"])
    writer.writeheader()
    writer.writerows(report_rows)
    with get_connection() as connection:
        connection.execute(
            """
            INSERT INTO student_import_reports(job_id, report_type, rows_json)
            VALUES (?, 'error', ?)
            """,
            (job_id, json.dumps(report_rows, ensure_ascii=False)),
        )
    return {
        "file_name": f"student_import_errors_{job_id}.csv",
        "content_type": "text/csv",
        "content": output.getvalue(),
        "total": len(report_rows),
    }
